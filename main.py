import tkinter as tk
from tkinter import ttk
from openpyxl import load_workbook, Workbook #biblioteca parqa trabalhar com arquivos Excel
from openpyxl.utils import get_column_letter #importa a função para converter números de colunas em letras
from datetime import date, datetime, timedelta
import matplotlib.pyplot as plt #biblioteca para criar gráficos
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg # importa a classe FigureCnvasTkAg  p exibir gráficos em Tkinter
import calendar

def salvar_valor():
    #obtem o valor digitado no campo de entrada e converte p float
    valor_dia = float(entry_valor.get())
    #add o valor a lista de valores
    valores.append(valor_dia)
    #calcula o total economizado
    total_valores = sum(valores)
    #limpa o campo de entrada
    entry_valor.delete(0, tk.END)
    
    label_status.config(text='Valor salvo com sucesso!', foreground='green')
    #atualiza o rótulo de total
    label_total.config(text=f'Total economizado: R${total_valores:.2f}')
    
    #adicionando valor em uma nova linha na planilha
    linha = len(valores) + 1
    coluna_data = get_column_letter(1)
    coluna_valor = get_column_letter(2)
    sheet.cell(row=linha, column=1, value=date.today().strftime("%d-%m-%y"))
    sheet.cell(row=linha, column=2, value=valor_dia)
    

def plotar_grafico():
    global FigureCanvasTkAgg
    
    #OBTER AS DATAS E OS VALORES DA PLANILHA
    datas = [cell.value.date() if isinstance(cell.value, datetime) else datetime.strptime(cell.value, '%d-%m-%y').date() for cell in sheet['A'][1:]]
    valores = [cell.value for cell in sheet['B'][1:]]
    
    #AGRUPAR OS VALORES POR MES atraves de um DICIONÁRIO
    dados_mensais = {}
    for data, valor in zip(datas, valores):
        mes_ano = data.strftime('%m-%Y')
        if mes_ano in dados_mensais:
            dados_mensais[mes_ano].append(valor)
        else:
            dados_mensais[mes_ano] = [valor]
            
    #cria um gráfico e barras do matplotlib 
    fig = plt.Figure(figsize=(12, 6), dpi=80)
    ax_barras = fig.add_subplot(121)
    ax_pie = fig.add_subplot(122)
    barras = ax_barras.bar(range(len(dados_mensais)), [sum(valores) for valores in dados_mensais.values()])

    for i, barra in enumerate(barras):
        altura = barra.get_height()
        ax_barras.text(barra.get_x() + barra.get_width() / 2, altura, f'R${altura:.2f}', ha='center', va='bottom')

    nomes_meses = []
    for mes_ano in dados_mensais.keys():
        mes, ano = mes_ano.split('-')
        nome_mes = calendar.month_name[int(mes)]
        nomes_meses.append(f'{nome_mes}-{ano}')

    ax_barras.set_xticks(range(len(dados_mensais)))
    ax_barras.set_xticklabels(nomes_meses, ha='right')

    ax_barras.spines['top'].set_visible(False)
    ax_barras.spines['right'].set_visible(False)
    ax_barras.spines['bottom'].set_visible(False)
    ax_barras.spines['left'].set_visible(False)

    ax_barras.set_title('Economia por Mês')
    ax_barras.title.set_position([.5, 8.05])
    ax_barras.set_xlabel('Mês')
    ax_barras.set_ylabel('Valor Economizado')

    data_inicial = min(datas)
    data_final = max(datas)
    diferenca = (data_final - data_inicial).days
    semanas = diferenca // 7

    labels = [f'{i+1}ª Semana' for i in range(semanas)]
    valores_semana = []
    for i in range(semanas):
        data_inicio = data_inicial + timedelta(weeks=i)
        data_fim = data_inicio + timedelta(weeks=1)
        valores_semana.append(sum(valor for data, valor in zip(datas, valores) if data_inicio <= data < data_fim))

    pie = ax_pie.pie(valores_semana, labels=labels, autopct='%1.1f%%', startangle=90)
    ax_pie.set_title('Economia por Semana')

    #destruindo o gráfico atual e plotando um novo
    canvas = FigureCanvasTkAgg(fig, master=window)
    canvas.get_tk_widget().pack(padx=10, pady=10)

    fig.tight_layout()
    
#CONFIGURANDO A INTERFACE GRÁFICA
window = tk.Tk() #cria uma instancia da classe Tk para criar uma nova janela
window.title('App de Poupança Pessoal')
window.geometry('700x500')
window.configure(bg='#252525')

style = ttk.Style() #cria uma instancia da classe Style() para personalizar os estilos dos widgets
style.theme_use('clam')
#TLabel define o estilo dos rotulos (label)
style.configure('TLabel', background= '#252525', foreground = '#FFFFFF', font=('Arial', 12))
#TEntry define o estilo dos campos de entrada
style.configure('TEntry', fieldbackground= '#FFFFFF', font=('Arial', 12))
#TButton define o estilo dos botões
style.configure('TButton', background= '#4CAF50', foreground = '#FFFFFF', font=('Arial', 12))

#cria um rótulo com o texto
label_instrucao = ttk.Label(window, text='Insira o valor diário')
#cria um rótulo vazio para exibir o status da operação
label_status = ttk.Label(window, text='', foreground='red')
#cria um rótulo para exibir o total economizado
label_total = ttk.Label(window, text='', font=('Arial', 14, 'bold'))
#cria um campo de entrada para oo usuario inserir o valor
entry_valor = ttk.Entry(window)

#cria um botão para salvar o valor
button_salvar = ttk.Button(window, text='Salvar', command=salvar_valor)

#POSICIONANDO OS ELEMENTOS NA JANELA

label_instrucao.pack(pady=10)
entry_valor.pack(pady=5)
button_salvar.pack(pady=10)
label_status.pack() #adiciona a label de status sem espaçamento
label_total.pack(pady=10) #adiciona um espaçamento de 10px acima e abaixo do rótulo de total

#CARREGAMENTO DA PLANILHA EXISTENTE OU CRIAÇÃO DE UMA NOVA

try:
    workbook = load_workbook('valores_diarios.xlsx')
except FileNotFoundError:
    workbook=Workbook()
    
#SELECIONANDO A PRIMEIRA PLANILHA
sheet = workbook.active

#VERIFICANDO SE A PLANILHA JÁ POSSUI VALORES SALVOS
if sheet.max_row == 0:
    sheet.cell(row=1, column=1, value='Data')
    sheet.cell(row=1, column=2, value='Valor diário')

#OBTEM A LISTA DE VALORES JÁ SALVOS
valores = [cell.value for cell in sheet['B'][1:]]

#Exibe o total economizado
label_total.config(text=f'Total economizado: R${sum(valores):.2f}')


#inicia o loop principal da aplicação
plotar_grafico()
window.mainloop()

#salva a planilha com os valores atualizados
try:
    workbook.save('valores_diarios.xlsx')
except PermissionError:
    print('Erro: Não foi possível salvar o arquivo!')
     